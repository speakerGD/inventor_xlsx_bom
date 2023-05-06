import sys
import os
import re
import openpyxl
import pprint
import operator
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from settings import *


def main():
    # Check that files are provided
    if len(sys.argv) != 4 or not all([arg.endswith(".xlsx") for arg in sys.argv[1:]]):
        sys.exit("Usage: bom_generator.py <source>.xlsx <template>.xlsx <result>.xlsx")

    # Check that provided inventor specification file exists
    if not os.path.exists("./" + sys.argv[1]):
        sys.exit(f"{sys.argv[1]} must be in the same folder with bom_generator.py")

    # Check that provided template file exists
    if not os.path.exists("./" + sys.argv[2]):
        sys.exit(f"{sys.argv[2]} must be in the same folder with bom_generator.py")

    # Load workbooks
    wb_source = openpyxl.load_workbook(sys.argv[1])
    wb_template = openpyxl.load_workbook(sys.argv[2])

    # Process bill of materials
    bill_of_materials(wb_source, wb_template)

    # Process bill of purchases parts
    bill_of_purchased(wb_source, wb_template)

    # Process bill of md1000 parts
    bill_of_md1000(wb_source, wb_template)

    # Save the result
    while True:
        match (input("Save the result? (y/n):")).lower():
            case "y" | "yes":
                wb_template.save(sys.argv[3])
                print(f"The result is saved in {sys.argv[3]}.")
                break
            case "n" | "no":
                print(f"The result has not been saved.")
                break
            case _:
                continue


def bill_of_materials(source: object, template: object) -> None:
    """
    Check the inventor .xlsx specification `source` for necessary columns to calculate BOM. Columns required:
    - bom_structure
    - quantity
    - material
    - mass
    Calculate materials from the `source` and copy data to the template.
    Delete unused rows of materials from the template.
    """

    columns = [
        PART_NUMBER,
        QUANTITY,
        MATERIAL,
        MASS,
        CUSTOM_MASS,
        CUSTOM_LENGTH,
        CUSTOM_WIDTH,
        CUSTOM_AREA,
        BOM_STRUCTURE,
    ]
    filters = {
        BOM_STRUCTURE: "^Обычный$",
    }

    # Active worksheet from the source
    sheet = source.active

    # Abort if not all columns on the sheet
    if missing := missing_columns(sheet, columns):
        print("Could not issue a bill of materials")
        print("Missing columns in the source file:")
        for column in missing:
            print(column)
        return

    print("Collecting data for materials...")

    # Collect material data from the source
    data = get_data(sheet, columns, filters)

    if not data:
        print("No material parts in the source file")
        return

    materials = get_materials_info(data)
    # Transfer data to the template
    try:
        sheet_profile_material = template[PROFILE_MATERIAL]
        sheet_flat_material = template[FLAT_MATERIAL]
    except KeyError:
        print(
            f"Template must contain '{PROFILE_MATERIAL}' and '{FLAT_MATERIAL}' sheets"
        )
        return
    else:
        print(f"Transferring data to the {PROFILE_MATERIAL} sheet...")
        transferred_materials = transfer_materials(sheet_profile_material, materials)

        print(f"Transferring data to the {FLAT_MATERIAL} sheet...")
        transferred_materials.update(transfer_materials(sheet_flat_material, materials))

        for material in transferred_materials:
            del materials[material]

        if materials:
            print("Couldn't transfer these materials:")
            pprint.pprint(materials)


def bill_of_purchased(source: object, template: object) -> None:
    """
    Check the inventor .xlsx specification `source` for necessary columns to derive a list of purchased parts. Columns required:
    - part_number
    - bom_structure
    - quantity
    Copy purchased parts from the `source` to the template.
    """

    columns = [
        PART_NUMBER,
        DESCRIPTION,
        PROJECT,
        VENDOR,
        QUANTITY,
        STOCK_NUMBER,
        BOM_STRUCTURE,
    ]
    filters = {BOM_STRUCTURE: "^Приобретенный$"}

    sheet = source.active

    if missing := missing_columns(sheet, columns):
        print("Could not issue a bill of purchased parts.")
        print("Missing columns in the source file:")
        for column in missing:
            print(column)

    print("Collecting data for purchased parts")

    # Collect purchased data from the source
    data = get_data(sheet, columns, filters)

    if not data:
        print("No purchased parts in the source file")
        return

    try:
        sheet = template[PURCHASED]
    except KeyError:
        print(f"Template must contain {PURCHASED} sheet")
        return
    else:
        print(f"Transferring data to the {PURCHASED} sheet...")
        transfer_purchased(sheet, data)


def bill_of_md1000(source: object, template: object) -> None:
    """
    Check the inventor .xlsx specification `source` for necessary columns to derive a list of md1000 parts. Columns required:
    - part_number
    - description
    - quantity
    Copy md1000 parts from the `source` to the template.
    """

    columns = [PART_NUMBER, DESCRIPTION, QUANTITY]
    filters = {PART_NUMBER: r"^МД1000\."}

    # Active worksheet from the source
    sheet = source.active

    # Abort in not all columns on the sheet
    if missing := missing_columns(sheet, columns):
        print("Could not issue a bill of md1000 parts")
        print("Missing columns in the source file:")
        for column in missing:
            print(column)
        return

    print("Collecting data for MD1000 parts")

    # Collect MD100 data from the source
    data = get_data(sheet, columns, filters)

    if not data:
        print("No MD1000 parts in the source file")
        return

    try:
        sheet = template[MD1000]
    except KeyError:
        print(f"Template must contain '{MD1000} sheet")
        return
    else:
        print(f"Transferring data to the {MD1000} sheet...")
        transfer_md1000(sheet, data)


def missing_columns(sheet: object, columns: list[str]) -> list:
    """
    Validate that all `columns` exist on the `sheet`.
    Return a list of missing columns.
    """
    # If no columns provided - no missing
    if not columns:
        return None

    missing = []

    for column in columns:
        # The first row with columns' titles
        if column not in [cell.value for cell in sheet[1]]:
            missing.append(column)

    return missing


def column_number(row: object, column: str) -> int:
    """
    Return the number of a `column` from the `row`.
    If `column` not in the `row` - return None.
    """
    for cell in row:
        if cell.value == column:
            return cell.column

    return None


def get_data(sheet: object, columns: list[str], filters: dict) -> list[dict]:
    """
    Retrieve data from the `sheet` from related `columns`.
    Choose only rows that comply with all `filters`.
    Numbers of columns and rows start from 1.
    """
    # Target columns and their numbers in the sheet
    columns = {column: column_number(sheet[1], column) for column in columns}

    # All the rows from the sheet, except the first row with columns titles
    rows = {i for i in range(2, sheet.max_row)}

    # First row with columns titles
    for title_cell in sheet[1]:
        # If a column is a filter
        if title_cell.value in filters:
            # Initialize a set of filtered rows
            filtered_rows = set()

            # Filter cells values in the column
            for data_cell in sheet[get_column_letter(title_cell.column)]:
                if matches := re.search(filters[title_cell.value], data_cell.value):
                    # Populate the set of numbers of filtered rows
                    filtered_rows.add(data_cell.row)

            # Keep only the rows that comply with a filter
            rows = rows.intersection(filtered_rows)

    # Collect data
    data = []
    for row in sorted(rows):
        data.append(
            {
                column: sheet.cell(row=row, column=columns[column]).value
                for column in columns
            }
        )

    # Replace none-type values with empty string values
    for row in data:
        for k, v in row.items():
            if not v:
                row[k] = ""

    return data


def get_materials_info(data: list[dict]) -> dict:
    """
    Get info for all materials in the `data`.
    Info includes mass and scope, and also type of the material.
    """
    # All unique materials from the data
    materials = {row[MATERIAL]: {} for row in data}

    # Calculate properties of all parts made of material
    for material, properties in materials.items():
        # Mass
        properties["mass"] = get_mass(material, data)
        # Scope
        properties["scope"] = get_scope(material, data)
        # Type in accordance with MATERIAL_TYPE
        properties["type"] = get_type(material)

    return materials


def transfer_materials(sheet: object, materials: dict) -> set:
    """
    Complete the `sheet` of the template with `materials`.
    """

    # Font to highlight cells in changed rows
    highlight = Font(color="FF0000")

    transferred_materials = set()

    # Column with short materials names in the template
    for cell in sheet["A"]:
        # Avoid cell with titles of groups of materials
        if cell.font.b:
            continue
        # Avoid empty cells
        if cell.value == None:
            continue

        # Current row
        i = cell.row

        # Type of material of the material in the cell
        material_type = sheet[f"B{i}"].value
        if not material_type:
            material_type = ""

        for material, properties in materials.items():
            # If cell's material among used materials
            if (
                material.lower().startswith(cell.value.lower())
                and properties["type"] in material_type
            ):
                # Transfer data
                sheet[f"D{i}"] = properties["mass"]
                sheet[f"G{i}"] = properties["scope"]

                # Highlight cells in the row
                for cell in sheet[i]:
                    cell.font = highlight

                # Update transferred materials
                transferred_materials.add(material)

    return transferred_materials


def get_mass(material: str, data: list[dict]) -> float:
    """
    Parse rows of data. Each row is a part.
    Each part has its own material, mass and quantity.
    """
    # Summ of masses of all parts made of the material
    summ_mass = 0

    for part in data:
        # Look for parts made of the material
        if part[MATERIAL] == material:
            try:
                # Quantity of the part
                n = int(part[QUANTITY])
            except ValueError:
                print(f"Invalid quantity for {part[PART_NUMBER]}")
                continue
            else:
                # Two options to derive mass
                mass = (part[CUSTOM_MASS], part[MASS])
                # First option is prefferable, but not always exists
                if mass[0]:
                    summ_mass += float(mass[0]) * n
                # Second option
                elif mass[1]:
                    summ_mass += float(mass[1].rstrip(" кг")) * n
                # No options for mass
                else:
                    print(f"Couldn't find mass for {part[PART_NUMBER]}")

    return summ_mass


def get_scope(material: str, data: list[dict]) -> float:
    """
    Parse rows of data. Each row is a part. Part is either profiled or flat.
    Scope stands for length of profiled parts and for area of flat parts.
    """

    # Summ of areas or lengths of all parts made of the material
    scope = 0

    for part in data:
        # Look for parts made of the material
        if part[MATERIAL] == material:
            try:
                # Quantity of the part
                n = int(part[QUANTITY])
            except ValueError:
                print(f"Invalid quantity for {part[PART_NUMBER]}")
                continue
            else:
                # If material is a flat material
                if material.startswith(FLAT_MATERIAL_PREFIX):
                    if area := part[CUSTOM_AREA]:
                        # Area in the source in mm2, scope in m2
                        scope += int(area) / 1000000
                    else:
                        print(f"Couldn't find area for {part[PART_NUMBER]}")
                else:
                    if length := part[CUSTOM_LENGTH]:
                        # Lwngth in the souce in mm, scope in m
                        scope += int(length) / 1000
                    else:
                        print(f"Couldn't find length for {part[PART_NUMBER]}")

    return scope


def get_type(material: str) -> str:
    """
    Return type of the material in accordance with MATERIAL_TYPE.
    """
    for k, v in MATERIAL_TYPE.items():
        for snippet in v:
            if snippet.lower() in material.lower():
                return k

    return ""


def transfer_md1000(sheet: object, data: list[dict]) -> None:
    """
    Transfer data of MD1000 parts to the `sheet` of the template.
    """

    for i, row in enumerate(sorted(data, key=lambda d: d[PART_NUMBER])):
        # Starting row in the template's sheet
        sheet_row = 3 + i

        sheet[f"A{sheet_row}"] = i + 1
        sheet[f"B{sheet_row}"] = row[PART_NUMBER]
        sheet[f"C{sheet_row}"] = row[DESCRIPTION]
        sheet[f"D{sheet_row}"] = row[QUANTITY]


def transfer_purchased(sheet: object, data: list[dict]) -> None:
    """
    Transfer data of purchased parts to the `sheet` of the remplate.
    """

    for i, row in enumerate(sorted(data, key=operator.itemgetter(VENDOR, PART_NUMBER))):
        # Row in the template's sheet
        sheet_row = 3 + i

        sheet[f"A{sheet_row}"] = i + 1

        if row[VENDOR]:
            sheet[f"B{sheet_row}"] = row[DESCRIPTION]
            sheet[f"C{sheet_row}"] = row[PROJECT]
            sheet[f"D{sheet_row}"] = row[VENDOR]
        else:
            sheet[f"B{sheet_row}"] = row[PART_NUMBER]

        sheet[f"E{sheet_row}"] = row[QUANTITY]
        sheet[f"F{sheet_row}"] = row[STOCK_NUMBER]


if __name__ == "__main__":
    main()
